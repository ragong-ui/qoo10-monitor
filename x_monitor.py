"""
X (Twitter) Qoo10 偽物モニタリング — Yahoo! リアルタイム検索 API 版
Yahoo! リアルタイム検索内部 API を使用 (認証不要・無料・日本語ツイート特化)
エンドポイント: https://search.yahoo.co.jp/realtime/api/v1/pagination
"""

import io, json, os, re, sys, time
from urllib.parse import quote
from datetime import datetime, timedelta
from pathlib import Path

from dotenv import load_dotenv

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

load_dotenv()

EMAIL_FROM     = os.getenv("EMAIL_FROM", "")
EMAIL_TO       = os.getenv("EMAIL_TO", "")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD", "")

BASE_DIR     = Path(__file__).parent
RESULTS_DIR  = BASE_DIR / "results"
LOGS_DIR     = BASE_DIR / "logs"
HISTORY_FILE = RESULTS_DIR / "x_url_history.json"

YAHOO_API_URL = "https://search.yahoo.co.jp/realtime/api/v1/pagination"
YAHOO_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Referer":         "https://search.yahoo.co.jp/realtime/search",
    "Accept":          "application/json, text/plain, */*",
    "Accept-Language": "ja,en-US;q=0.9",
}

_FRAUD_WORDS = [
    "偽物", "ニセモノ", "にせもの", "パチモン", "パチもん", "パチモノ",
    "コピー", "fake", "偽造品", "模倣品", "コピー商品", "模造品", "詐欺",
    "中国",
]

EXCLUDE_KEYWORDS = [
    "見分け方", "見分け方法",
    "#PR", "#広告", "#ad", "#sponsored", "#タイアップ", "#案件",
    "弊社が判断した場合",
]

X_QUERIES = [
    # Qoo10 直接関連 (13種)
    "Qoo10 偽物",
    "Qoo10 ニセモノ",
    "Qoo10 にせもの",
    "Qoo10 パチモン",
    "Qoo10 パチもん",
    "Qoo10 パチモノ",
    "Qoo10 コピー",
    "Qoo10 fake",
    "Qoo10 偽造品",
    "Qoo10 模倣品",
    "Qoo10 コピー商品",
    "Qoo10 模造品",
    "Qoo10 詐欺",
    # メガ割 関連 (2種) — 偽物系KWフィルタを処理ループで適用
    "メガ割り",
    "メガ割",
    # 規制・被害対応 (3種)
    "Qoo10 販売禁止商品",
    "Qoo10 規約違反",
    "Qoo10 強制返金",
]

OFFICIAL_SCREEN_NAMES = {
    "qoo10fashion", "qoo10_shopping", "qoo10jp", "qoo10",
}

QOO10_URL_RE    = re.compile(r"https?://(?:www\.)?qoo10\.jp/\S+")
QOO10_DOMAIN_RE = re.compile(r"qoo10\.jp")


# ── URL 履歴管理 ────────────────────────────────────────────
def load_url_history() -> set:
    if not HISTORY_FILE.exists():
        return set()
    data = json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
    cutoff = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    return {url for url, date in data.items() if date >= cutoff}


def save_url_history(new_urls: dict[str, str]):
    existing: dict = {}
    if HISTORY_FILE.exists():
        existing = json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
    existing.update(new_urls)
    cutoff = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    pruned = {url: date for url, date in existing.items() if date >= cutoff}
    HISTORY_FILE.write_text(json.dumps(pruned, ensure_ascii=False, indent=2), encoding="utf-8")


def proximity_ok(text: str, query: str, fraud_words: list[str], window: int = 100) -> bool:
    """クエリ語(Qoo10 / メガ割)とfake系ワードがwindow文字以内に共存するか確認。"""
    anchors = ["Qoo10"] if "Qoo10" in query else [query.split()[0], "Qoo10"]
    for anchor in anchors:
        pos = 0
        while True:
            idx = text.find(anchor, pos)
            if idx == -1:
                break
            nearby = text[max(0, idx - window): idx + len(anchor) + window]
            if any(fw in nearby for fw in fraud_words):
                return True
            pos = idx + 1
    return False


# ── Yahoo! リアルタイム検索 API ─────────────────────────────
_HIGHLIGHT_RE = re.compile(r"\tSTART\t|\tEND\t")


def _extract_tweets(data: dict) -> list[dict]:
    """レスポンス JSON からツイートリストを取り出す
    構造: data['timeline']['entry'] (list of tweet dicts)
    """
    timeline = data.get("timeline")
    if isinstance(timeline, dict):
        entry = timeline.get("entry")
        if isinstance(entry, list):
            return entry
    return []


def _parse_tweet(raw: dict) -> dict | None:
    """ツイートオブジェクトを正規化
    フィールド: url / displayText / screenName (トップレベル)
    """
    try:
        url         = raw.get("url") or ""
        screen_name = (raw.get("screenName") or "").lower()
        # \tSTART\t…\tEND\t ハイライトマーカーを除去
        text = _HIGHLIGHT_RE.sub("", raw.get("displayText") or raw.get("displayTextBody") or "")
        return {"url": url, "text": text, "screen_name": screen_name}
    except Exception:
        return None


def search_x_yahoo(query: str) -> list[dict]:
    """Yahoo! リアルタイム検索 API で直近2日のツイートを40件取得"""
    import requests

    since_ts = int((datetime.now() - timedelta(days=2)).timestamp())
    params = {
        "p":       query,
        "results": 40,
        "since":   since_ts,
    }
    try:
        resp = requests.get(
            YAHOO_API_URL, params=params, headers=YAHOO_HEADERS, timeout=15
        )
        resp.raise_for_status()
        data = resp.json()
        return _extract_tweets(data)
    except Exception as e:
        print(f"  [ERROR] Yahoo API ({query}): {e}")
        return []


# ── メイン処理 ──────────────────────────────────────────────
def run_x_searches() -> list[dict]:
    date_str  = datetime.now().strftime("%Y-%m-%d")
    rows: list[dict] = []
    seen_urls  = load_url_history()
    new_urls: dict[str, str] = {}

    for i, query in enumerate(X_QUERIES):
        print(f"  Searching X (Yahoo): {query}")
        raw_tweets = search_x_yahoo(query)

        found = 0
        for raw in raw_tweets:
            parsed = _parse_tweet(raw)
            if not parsed:
                continue

            url         = parsed["url"].split("?")[0]  # utm パラメータ除去
            text        = parsed["text"]
            screen_name = parsed["screen_name"]

            if not url or url in seen_urls:
                continue

            # x.com/status/ のツイートのみ対象
            if "/status/" not in url:
                continue

            # 公式アカウント除外
            if screen_name in OFFICIAL_SCREEN_NAMES:
                continue

            # 広告・ガイド系 하드 제외
            if any(w in text for w in EXCLUDE_KEYWORDS):
                continue

            # クエリ語とfake系ワードが100文字以内に共存するか確認
            if not proximity_ok(text, query, _FRAUD_WORDS):
                continue

            has_qoo10_url = bool(QOO10_URL_RE.search(text))
            has_qoo10_ref = bool(QOO10_DOMAIN_RE.search(text))
            likelihood    = "HIGH" if has_qoo10_url else "MEDIUM"

            seen_urls.add(url)
            new_urls[url] = date_str
            rows.append({
                "date":       date_str,
                "query":      query,
                "url":        url,
                "summary":    text[:400],
                "qoo10_link": (
                    QOO10_URL_RE.search(text).group(0) if has_qoo10_url
                    else ("qoo10.jp 언급 있음" if has_qoo10_ref else "なし")
                ),
                "likelihood": likelihood,
                "search_url": "https://search.yahoo.co.jp/realtime/search?p=" + quote(query),
            })
            found += 1

        print(f"  → {found}件")

        # API への過負荷を避けるため少し待機
        if i < len(X_QUERIES) - 1:
            time.sleep(1)

    RESULTS_DIR.mkdir(exist_ok=True)
    save_url_history(new_urls)
    rows.sort(key=lambda x: 0 if x["likelihood"] == "HIGH" else 1)
    return rows


# ── 번역 유틸리티 ────────────────────────────────────────────
def _translate_batch_ja_ko(texts: list[str]) -> list[str]:
    """일본어 텍스트를 한국어로 배치 번역. 실패 시 원문 반환."""
    if not texts:
        return []
    try:
        from deep_translator import GoogleTranslator
        translator = GoogleTranslator(source="ja", target="ko")
        results = []
        for i in range(0, len(texts), 50):
            chunk = [t[:500] for t in texts[i:i+50]]
            translated = translator.translate_batch(chunk)
            results.extend(translated if translated else chunk)
        return results
    except Exception as e:
        print(f"  [WARN] 번역 실패: {e}")
        return texts


def _bilingual(jp: str, kr: str) -> str:
    """일본어+한국어 병기 셀 내용 생성"""
    kr = (kr or "").strip()
    if kr and kr != jp.strip():
        return f"{jp}\n{'─' * 22}\n[한국어]\n{kr}"
    return jp


# ── Excel 出力 ──────────────────────────────────────────────
def save_excel(rows: list[dict], date_str: str) -> Path:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    RESULTS_DIR.mkdir(exist_ok=True)
    path = RESULTS_DIR / f"X_Qoo10_偽物レポート_{date_str}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "X게시물 레포트"

    headers = ["검색일 / 検索日", "검색 쿼리 / クエリ", "게시물 URL / 投稿URL", "게시물 내용 / 投稿内容", "Qoo10 상품 URL", "위험도 / 危険度", "검색확인 / 検索確認", "오탐지여부", "Status", "아카이브 / Archive"]
    ws.append(headers)

    hdr_fill  = PatternFill("solid", fgColor="2F5496")
    high_fill = PatternFill("solid", fgColor="FFD7D7")
    med_fill  = PatternFill("solid", fgColor="FFF3CD")

    for c in ws[1]:
        c.font      = Font(bold=True, color="FFFFFF")
        c.fill      = hdr_fill
        c.alignment = Alignment(horizontal="center")

    print("  Translating summaries (JP→KR)...")
    kr_list = _translate_batch_ja_ko([r["summary"] for r in rows])

    for r, kr in zip(rows, kr_list):
        ws.append([
            r["date"], r["query"], r["url"],
            _bilingual(r["summary"], kr), r["qoo10_link"], r["likelihood"],
            "", "", "New",   # G=검색확인(placeholder), H=오탐지여부, I=Status
        ])
        fill = high_fill if r["likelihood"] == "HIGH" else med_fill
        for c in ws[ws.max_row]:
            c.fill      = fill
            c.alignment = Alignment(wrap_text=True, vertical="top")
        # G열: Yahoo! 리얼타임 검색결과 링크
        if r.get("search_url"):
            g_cell = ws.cell(row=ws.max_row, column=7)
            g_cell.value = "Yahoo検索"
            g_cell.hyperlink = r["search_url"]
            g_cell.font = Font(bold=False, color="0563C1", underline="single")
            g_cell.fill = fill
        # J열: Wayback Machine 아카이브 링크
        j_cell = ws.cell(row=ws.max_row, column=10)
        j_cell.value = "Wayback"
        j_cell.hyperlink = f"https://web.archive.org/web/{r['url']}"
        j_cell.font = Font(bold=False, color="0563C1", underline="single")
        j_cell.fill = fill

    for col, w in zip("ABCDEFGHIJ", [12, 20, 55, 80, 50, 10, 12, 12, 14, 10]):
        ws.column_dimensions[col].width = w

    from openpyxl.comments import Comment
    ws["H1"].comment = Comment("O = 오탐지 (오검지)\nX = 실검지 (위조품 확인)\n공백 = 미확인", "x_monitor")

    # H열 O/X 드롭다운 (오탐지여부)
    dv_ox = DataValidation(type="list", formula1='"O,X"', allow_blank=True)
    ws.add_data_validation(dv_ox)
    dv_ox.add("H2:H10000")

    # I열 Status 드롭다운
    dv_status = DataValidation(type="list", formula1='"New,Reviewing,Actioned,Closed"', allow_blank=False)
    ws.add_data_validation(dv_status)
    dv_status.add("I2:I10000")

    ws.freeze_panes = "A2"

    wb.save(path)
    print(f"  Excel saved: {path}")
    return path, kr_list


# ── Gmail 送信 ──────────────────────────────────────────────
def send_email(path: Path, summary: str, date_str: str):
    import smtplib
    from email import encoders
    from email.mime.base import MIMEBase
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    if not EMAIL_PASSWORD:
        print("  [SKIP] EMAIL_PASSWORD not set.")
        return

    try:
        msg            = MIMEMultipart()
        msg["From"]    = EMAIL_FROM
        msg["To"]      = EMAIL_TO
        msg["Subject"] = f"[X 위조품 모니터링 / 偽物モニタリング] {date_str} 레포트"
        body = (
            f"X (Twitter) Qoo10 위조품 모니터링 레포트입니다 / 偽物モニタリングレポートです。\n\n"
            f"{summary}\n\n"
            f"첨부 Excel을 확인해 주세요 / 添付のExcelをご確認ください。\n\n"
            f"※ HIGH   = 위조품 키워드 + Qoo10 상품 URL 둘 다 확인 / 偽物系KW + Qoo10商品URL 両方確認\n"
            f"※ MEDIUM = 위조품 키워드만 (Qoo10 URL 없음) / 偽物系KWのみ（Qoo10 URLなし）\n\n"
            f"[취득원: Yahoo! 리얼타임 검색 API / 직근 2일분 / Yahoo! リアルタイム検索 API / 直近2日分]"
        )
        msg.attach(MIMEText(body, "plain", "utf-8"))

        with open(path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename={path.name}")
            msg.attach(part)

        with smtplib.SMTP("smtp.gmail.com", 587) as s:
            s.ehlo(); s.starttls()
            s.login(EMAIL_FROM, EMAIL_PASSWORD)
            s.send_message(msg)
        print("  Email sent via Gmail.")
    except Exception as e:
        print(f"  [ERROR] Email: {e}")


def send_no_issue_email(date_str: str):
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    if not EMAIL_PASSWORD:
        print("  [SKIP] EMAIL_PASSWORD not set.")
        return

    try:
        msg            = MIMEMultipart()
        msg["From"]    = EMAIL_FROM
        msg["To"]      = EMAIL_TO
        msg["Subject"] = f"[X 위조품 모니터링 / 偽物モニタリング] {date_str} 이상없음 / 異常なし"
        msg.attach(MIMEText(
            "오늘의 X Qoo10 위조품 모니터링 결과, 새로운 검지가 없었습니다 / 本日のX Qoo10偽物モニタリングの結果、新たな検知はありませんでした。\n\n"
            "이상없음 / 異常なし ✅\n\n[취득원: Yahoo! 리얼타임 검색 API / 직근 2일분 / Yahoo! リアルタイム検索 API / 直近2日分]",
            "plain", "utf-8"
        ))
        with smtplib.SMTP("smtp.gmail.com", 587) as s:
            s.ehlo(); s.starttls()
            s.login(EMAIL_FROM, EMAIL_PASSWORD)
            s.send_message(msg)
        print("  Email sent (異常なし).")
    except Exception as e:
        print(f"  [ERROR] Email: {e}")


# ── エントリポイント ────────────────────────────────────────
def main():
    date_str = datetime.now().strftime("%Y-%m-%d")
    LOGS_DIR.mkdir(exist_ok=True)

    print(f"[{datetime.now():%Y-%m-%d %H:%M}] X Qoo10 偽物モニタリング 開始")
    print(f"  方式: Yahoo! リアルタイム検索 API / クエリ数: {len(X_QUERIES)} / 直近2日")

    rows = run_x_searches()

    high    = sum(1 for r in rows if r["likelihood"] == "HIGH")
    med     = sum(1 for r in rows if r["likelihood"] == "MEDIUM")
    summary = (
        f"  ・危険度 HIGH  (偽物KW + Qoo10商品URL): {high} 件\n"
        f"  ・危険度 MEDIUM (偽物KWのみ)           : {med} 件\n"
        f"  ・合計                                : {len(rows)} 件"
    )
    print(summary)

    if rows:
        path, kr_list = save_excel(rows, date_str)
        send_email(path, summary, date_str)
        try:
            from gsheets import write_to_sheets
            write_to_sheets(rows, kr_list, "x")
        except Exception as e:
            print(f"  [WARN] Google Sheets 기록 실패 (Excel은 정상 저장됨): {e}")
    else:
        print("  検知なし")
        send_no_issue_email(date_str)

    print(f"[{datetime.now():%Y-%m-%d %H:%M}] 完了")


if __name__ == "__main__":
    main()
