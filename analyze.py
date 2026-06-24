"""
Qoo10 偽物モニタリング — 주간 분석 스크립트
Usage: python analyze.py

담당자가 Excel의 '오탐지여부' 열에 입력한 결과를 읽어
EXCLUDE_KEYWORDS / REVIEW_SIGNALS 갱신 후보를 출력한다.

오탐지여부 입력 규칙:
  O  → 오탐지 (오검지 — 광고·가이드·무관 페이지 등)
  X  → 실검지 (위조품 리뷰 확인)
  공백 → 미확인 (분석 제외)
"""

import re
import sys
import io
import openpyxl
from collections import Counter, defaultdict
from pathlib import Path
from urllib.parse import urlparse

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

RESULTS_DIR = Path(__file__).parent / "results"

# ── monitor.py 와 동기화 필요 ──────────────────────────────
CURRENT_EXCLUDE = [
    "見分け方", "見分け方法",
    "#PR", "#広告", "#ad", "#sponsored", "#タイアップ", "#案件",
]
CURRENT_REVIEW_SIGNALS = [
    "買った", "届いた", "購入した", "買ってしまった",
    "偽物が届いた", "ニセモノが届いた", "買わされた",
    "つかまされた", "やられた", "体験談", "実際に",
    "買ってみた", "被害", "詐欺にあった", "強制返金", "規約違反",
    "bought", "received", "purchased", "got a fake",
]
CURRENT_GUIDE_SIGNALS = [
    "見分け方", "見分け方法", "ガイド", "まとめ", "how to spot",
    "complete guide", "チェックポイント", "対策方法",
    "見極め", "注意点", "安全な買い方",
]

# 오탐지여부 열: O = 오탐지(NG), X = 실검지(OK) — 구 확인결과 열과 반전
_FALSPOS_TRUE  = {"o", "〇", "○", "ok"}   # 오탐지
_FALSPOS_FALSE = {"x", "×", "✗", "ng"}   # 실검지

# 구 확인결과 열 호환 (레거시 파일용)
_OK_LEGACY = {"〇", "○", "o", "ok", "✅", "본물", "confirmed"}
_NG_LEGACY = {"×", "✗", "x", "ng", "❌", "오검지", "false", "광고", "ad"}


def normalize_confirm(val: str, is_falspos_col: bool = False) -> str:
    """오탐지여부 열(is_falspos_col=True): O→NG, X→OK  /  구 확인결과 열: O→OK, X→NG"""
    v = str(val or "").strip().lower()
    if is_falspos_col:
        if v in _FALSPOS_TRUE:
            return "NG"
        if v in _FALSPOS_FALSE:
            return "OK"
    else:
        if v in _OK_LEGACY:
            return "OK"
        if v in _NG_LEGACY:
            return "NG"
    return ""


def load_reviewed_rows() -> tuple[list[dict], list[dict]]:
    """results/ 폴더의 모든 Excel에서 확인 완료 행을 로드."""
    rows_ok, rows_ng = [], []

    xlsx_files = sorted(RESULTS_DIR.glob("Qoo10_偽物レポート_*.xlsx"))
    if not xlsx_files:
        print("[ERROR] results/ 폴더에 레포트 Excel이 없습니다.")
        return [], []

    for xlsx in xlsx_files:
        try:
            wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
        except Exception as e:
            print(f"[WARN] {xlsx.name} 읽기 실패: {e}")
            continue

        # 시트 찾기 (한국어/일본어 시트명 모두 대응)
        sheet = None
        for name in wb.sheetnames:
            if "레포트" in name or "レポート" in name:
                sheet = wb[name]
                break
        if sheet is None:
            continue

        rows_iter = sheet.iter_rows(values_only=True)
        headers = [str(c or "").strip() for c in next(rows_iter, [])]

        # 열 인덱스 탐색
        def find_col(candidates):
            for c in candidates:
                for i, h in enumerate(headers):
                    if c in h:
                        return i
            return None

        confirm_idx = find_col(["오탐지여부", "확인결과", "確認結果"])
        summary_idx = find_col(["개요", "概要"])
        kw_idx      = find_col(["키워드", "キーワード"])
        url_idx     = find_col(["URL"])
        level_idx   = find_col(["위험도", "危険度"])

        if confirm_idx is None:
            continue  # 확인결과 열 없는 파일은 스킵

        is_falspos_col = confirm_idx < len(headers) and "오탐지여부" in headers[confirm_idx]

        for row in rows_iter:
            confirm = normalize_confirm(
                row[confirm_idx] if confirm_idx < len(row) else "",
                is_falspos_col=is_falspos_col,
            )
            if not confirm:
                continue

            entry = {
                "file":     xlsx.name,
                "keyword":  row[kw_idx]  if kw_idx  is not None and kw_idx  < len(row) else "",
                "url":      row[url_idx] if url_idx  is not None and url_idx  < len(row) else "",
                "summary":  str(row[summary_idx] if summary_idx is not None and summary_idx < len(row) else "") or "",
                "level":    row[level_idx] if level_idx is not None and level_idx < len(row) else "",
            }
            (rows_ok if confirm == "OK" else rows_ng).append(entry)

    return rows_ok, rows_ng


def extract_words(text: str) -> list[str]:
    """스니펫에서 의미 있는 어절/단어 추출 (MeCab 없이)."""
    words = []
    # 카타카나 3자 이상
    words += re.findall(r'[ァ-ヶー]{3,}', text)
    # 한자 2자 이상
    words += re.findall(r'[一-龯々]{2,}', text)
    # 히라가나 3자 이상
    words += re.findall(r'[ぁ-ん]{3,}', text)
    # 해시태그
    words += re.findall(r'#\S+', text)
    # 영문 3자 이상 (URL 제외)
    words += [w for w in re.findall(r'[a-zA-Z]{3,}', text)
              if w.lower() not in {"http", "https", "www", "com", "the", "and", "for"}]
    return words


def score_ng_candidates(rows_ok: list[dict], rows_ng: list[dict]) -> list[tuple[str, int, int, float]]:
    """× 행에 자주 나오고 〇 행에 적게 나오는 단어 후보를 점수 순으로 반환."""
    ng_words = Counter()
    ok_words = Counter()

    for r in rows_ng:
        for w in set(extract_words(r["summary"])):
            ng_words[w] += 1
    for r in rows_ok:
        for w in set(extract_words(r["summary"])):
            ok_words[w] += 1

    all_words = set(ng_words) | set(ok_words)
    scored = []
    for w in all_words:
        ng_c = ng_words[w]
        ok_c = ok_words[w]
        if ng_c < 2:
            continue
        if w in CURRENT_EXCLUDE + CURRENT_GUIDE_SIGNALS:
            continue
        # 오검지 출현 비율이 실검지의 3배 이상인 것만
        ratio = ng_c / (ok_c + 0.5)
        if ratio >= 3.0:
            scored.append((w, ng_c, ok_c, ratio))

    return sorted(scored, key=lambda x: (-x[1], -x[3]))


def analyze():
    print("=" * 60)
    print("  Qoo10 モニタリング 주간 분석 리포트")
    print("=" * 60)

    rows_ok, rows_ng = load_reviewed_rows()
    total = len(rows_ok) + len(rows_ng)

    if total == 0:
        print("\n확인된 데이터가 없습니다.")
        print("Excel의 '오탐지여부' 열에 O (오탐지) 또는 X (실검지) 를 입력해주세요.")
        return

    print(f"\n[확인 완료] 총 {total}건  |  X 실검지 (위조품): {len(rows_ok)}건  |  O 오탐지: {len(rows_ng)}건")
    print(f"정확도: {len(rows_ok)/total*100:.1f}%  |  오검지율: {len(rows_ng)/total*100:.1f}%\n")

    # ── 오검지 URL 도메인 분포 ──────────────────────────────
    if rows_ng:
        print("─" * 60)
        print("【오검지(×) URL 도메인 분포】")
        domains = Counter(urlparse(str(r["url"])).netloc for r in rows_ng)
        for domain, cnt in domains.most_common(10):
            print(f"  {domain:<45} {cnt}건")

    # ── 오검지 키워드 분포 ──────────────────────────────────
    if rows_ng:
        print("\n【오검지(×) 검색 키워드 분포】")
        kw_ng = Counter(str(r["keyword"]) for r in rows_ng)
        for kw, cnt in kw_ng.most_common():
            print(f"  {kw:<35} {cnt}건")

    # ── EXCLUDE_KEYWORDS 추가 후보 ─────────────────────────
    print("\n─" * 60)
    print("【EXCLUDE_KEYWORDS 추가 후보 (오검지에 자주 등장)】")
    candidates = score_ng_candidates(rows_ok, rows_ng)
    if candidates:
        print(f"  {'단어':<20} {'오검지':<8} {'실검지':<8} {'비율'}")
        for word, ng_c, ok_c, ratio in candidates[:20]:
            print(f"  {word:<20} {ng_c}건{'':<5} {ok_c}건{'':<5} ×{ratio:.1f}")
    else:
        print("  → 새로운 후보 없음 (현재 필터가 적절히 작동 중)")

    # ── 오검지 행 상세 (검토용) ────────────────────────────
    if rows_ng:
        print("\n─" * 60)
        print("【오검지(×) 상세 목록】")
        by_kw = defaultdict(list)
        for r in rows_ng:
            by_kw[str(r["keyword"])].append(r)
        for kw, entries in sorted(by_kw.items()):
            print(f"\n  ▶ {kw}")
            for e in entries:
                print(f"    URL    : {e['url']}")
                summary_preview = str(e["summary"])[:120].replace("\n", " ")
                print(f"    개요   : {summary_preview}")

    # ── 실검지 분포 ────────────────────────────────────────
    if rows_ok:
        print("\n─" * 60)
        print("【실검지(〇) 검색 키워드 분포】")
        kw_ok = Counter(str(r["keyword"]) for r in rows_ok)
        for kw, cnt in kw_ok.most_common():
            print(f"  {kw:<35} {cnt}건")

        print("\n【실검지(〇) 자주 등장하는 표현 → REVIEW_SIGNALS 강화 후보】")
        ok_words = Counter()
        for r in rows_ok:
            for w in set(extract_words(r["summary"])):
                if w not in CURRENT_REVIEW_SIGNALS:
                    ok_words[w] += 1
        for word, cnt in ok_words.most_common(15):
            if cnt >= 2:
                print(f"  '{word}' — {cnt}건")

    # ── 최종 추천 요약 ─────────────────────────────────────
    print("\n" + "=" * 60)
    print("【Claude에게 전달할 추천 요약】")
    print("=" * 60)
    if candidates:
        top = [w for w, *_ in candidates[:10]]
        print(f"EXCLUDE_KEYWORDS 추가 후보: {top}")
    else:
        print("EXCLUDE_KEYWORDS: 변경 불필요")

    ok_top = []
    if rows_ok:
        ok_words_flat = Counter()
        for r in rows_ok:
            for w in set(extract_words(r["summary"])):
                if w not in CURRENT_REVIEW_SIGNALS and len(w) >= 3:
                    ok_words_flat[w] += 1
        ok_top = [w for w, c in ok_words_flat.most_common(10) if c >= 2]
    if ok_top:
        print(f"REVIEW_SIGNALS 추가 후보:  {ok_top}")
    else:
        print("REVIEW_SIGNALS: 변경 불필요")

    print("\n→ 이 리포트를 Claude에게 공유하면 키워드 업데이트를 제안받을 수 있습니다.")


# ════════════════════════════════════════════════════════════
# X (Twitter) 모니터링 분석
# ════════════════════════════════════════════════════════════

# ── x_monitor.py 와 동기화 필요 ───────────────────────────
X_CURRENT_FRAUD_WORDS = [
    "偽物", "ニセモノ", "にせもの", "パチモン", "パチもん", "パチモノ",
    "コピー", "fake", "偽造品", "模倣品", "コピー商品", "模造品", "詐欺",
]
X_OFFICIAL_ACCOUNTS = {
    "qoo10fashion", "qoo10_shopping", "qoo10jp", "qoo10",
}


def load_x_reviewed_rows() -> tuple[list[dict], list[dict]]:
    """X_Qoo10_偽物レポート_*.xlsx 에서 확인 완료 행을 로드."""
    rows_ok, rows_ng = [], []

    xlsx_files = sorted(RESULTS_DIR.glob("X_Qoo10_偽物レポート_*.xlsx"))
    if not xlsx_files:
        return [], []

    for xlsx in xlsx_files:
        try:
            wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
        except Exception as e:
            print(f"[WARN] {xlsx.name} 읽기 실패: {e}")
            continue

        sheet = None
        for name in wb.sheetnames:
            if "레포트" in name or "レポート" in name:
                sheet = wb[name]
                break
        if sheet is None:
            continue

        rows_iter = sheet.iter_rows(values_only=True)
        headers = [str(c or "").strip() for c in next(rows_iter, [])]

        def find_col(candidates):
            for c in candidates:
                for i, h in enumerate(headers):
                    if c in h:
                        return i
            return None

        confirm_idx = find_col(["오탐지여부", "확인결과", "確認結果"])
        summary_idx = find_col(["내용", "投稿内容"])
        query_idx   = find_col(["쿼리", "クエリ"])
        url_idx     = find_col(["URL"])
        level_idx   = find_col(["위험도", "危険度"])

        if confirm_idx is None:
            continue

        is_falspos_col = confirm_idx < len(headers) and "오탐지여부" in headers[confirm_idx]

        for row in rows_iter:
            confirm = normalize_confirm(
                row[confirm_idx] if confirm_idx < len(row) else "",
                is_falspos_col=is_falspos_col,
            )
            if not confirm:
                continue
            entry = {
                "file":    xlsx.name,
                "query":   row[query_idx]   if query_idx   is not None and query_idx   < len(row) else "",
                "url":     row[url_idx]     if url_idx     is not None and url_idx     < len(row) else "",
                "summary": str(row[summary_idx] if summary_idx is not None and summary_idx < len(row) else "") or "",
                "level":   row[level_idx]   if level_idx   is not None and level_idx   < len(row) else "",
            }
            (rows_ok if confirm == "OK" else rows_ng).append(entry)

    return rows_ok, rows_ng


def analyze_x():
    print("\n" + "=" * 60)
    print("  X (Twitter) 모니터링 주간 분석 리포트")
    print("=" * 60)

    rows_ok, rows_ng = load_x_reviewed_rows()
    total = len(rows_ok) + len(rows_ng)

    if total == 0:
        print("\nX 모니터링 확인된 데이터가 없습니다.")
        print("X_Qoo10_偽物レポート_*.xlsx 의 '오탐지여부' 열에 O (오탐지) 또는 X (실검지) 를 입력해주세요.")
        return

    print(f"\n[확인 완료] 총 {total}건  |  X 실검지 (위조품): {len(rows_ok)}건  |  O 오탐지: {len(rows_ng)}건")
    print(f"정확도: {len(rows_ok)/total*100:.1f}%  |  오검지율: {len(rows_ng)/total*100:.1f}%\n")

    # ── 오검지 계정 분포 ───────────────────────────────────
    if rows_ng:
        print("─" * 60)
        print("【오검지(×) URL에서 자주 나오는 계정 패턴】")
        account_re = re.compile(r"x\.com/([^/]+)/status")
        ng_accounts = Counter()
        for r in rows_ng:
            m = account_re.search(str(r["url"]))
            if m:
                ng_accounts[m.group(1).lower()] += 1
        for acc, cnt in ng_accounts.most_common(10):
            flag = " ← 공식계정 추가 검토" if acc not in X_OFFICIAL_ACCOUNTS else " (이미 공식)"
            print(f"  @{acc:<40} {cnt}건{flag}")

    # ── 오검지 쿼리 분포 ───────────────────────────────────
    if rows_ng:
        print("\n【오검지(×) 검색 쿼리 분포】")
        ng_queries = Counter(str(r["query"]) for r in rows_ng)
        for q, cnt in ng_queries.most_common():
            print(f"  {q:<35} {cnt}건")

    # ── 오검지 텍스트 패턴 → 제외 후보 ───────────────────
    print("\n─" * 60)
    print("【오검지(×) 게시물에 자주 등장하는 표현 → 제외 후보】")
    ng_candidates = score_ng_candidates(rows_ok, rows_ng)
    if ng_candidates:
        print(f"  {'단어':<20} {'오검지':<8} {'실검지':<8} {'비율'}")
        for word, ng_c, ok_c, ratio in ng_candidates[:15]:
            print(f"  {word:<20} {ng_c}건{'':<5} {ok_c}건{'':<5} ×{ratio:.1f}")
    else:
        print("  → 새로운 후보 없음")

    # ── 오검지 상세 ────────────────────────────────────────
    if rows_ng:
        print("\n【오검지(×) 상세 목록】")
        by_q = defaultdict(list)
        for r in rows_ng:
            by_q[str(r["query"])].append(r)
        for q, entries in sorted(by_q.items()):
            print(f"\n  ▶ {q}")
            for e in entries:
                print(f"    URL    : {e['url']}")
                print(f"    내용   : {str(e['summary'])[:100].replace(chr(10), ' ')}")

    # ── 실검지 분포 ────────────────────────────────────────
    if rows_ok:
        print("\n─" * 60)
        print("【실검지(〇) 검색 쿼리 분포】")
        ok_queries = Counter(str(r["query"]) for r in rows_ok)
        for q, cnt in ok_queries.most_common():
            print(f"  {q:<35} {cnt}건")

        print("\n【실검지(〇) 자주 등장하는 표현 → FRAUD_WORDS 강화 후보】")
        ok_words = Counter()
        for r in rows_ok:
            for w in set(extract_words(r["summary"])):
                if w not in X_CURRENT_FRAUD_WORDS and len(w) >= 3:
                    ok_words[w] += 1
        for word, cnt in ok_words.most_common(15):
            if cnt >= 2:
                print(f"  '{word}' — {cnt}건")

    # ── 최종 추천 요약 ────────────────────────────────────
    print("\n" + "=" * 60)
    print("【X 분석 — Claude에게 전달할 추천 요약】")
    print("=" * 60)

    if ng_candidates:
        top = [w for w, *_ in ng_candidates[:10]]
        print(f"X 오검지 제외 후보 표현: {top}")
    else:
        print("X 제외 후보: 없음")

    if rows_ng:
        ng_acc_top = [acc for acc, _ in ng_accounts.most_common(5)
                      if acc not in X_OFFICIAL_ACCOUNTS]
        if ng_acc_top:
            print(f"OFFICIAL_SCREEN_NAMES 추가 후보 계정: {ng_acc_top}")

    print("\n→ 이 리포트를 Claude에게 공유하면 x_monitor.py 키워드 업데이트를 제안받을 수 있습니다.")


if __name__ == "__main__":
    analyze()
    analyze_x()
