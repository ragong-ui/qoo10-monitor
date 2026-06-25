"""
Qoo10 偽物モニタリング — 毎日実行スクリプト
・SerpAPI で Google 検索（過去2日）
・疑わしいページを Excel に出力
・Gmail でメール送信
・Slack Incoming Webhook で通知
"""

import io
import json
import os
import re
import sys
import requests
import openpyxl
from collections import Counter
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import quote

from dotenv import load_dotenv

# Windows CP949 환경에서 UTF-8 출력 강제
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

load_dotenv()

# ── 検索キーワード ──────────────────────────────────────
# 文字列: そのまま検索・そのまま検知
# dict : {"query": 検索語, "and_any": [追加条件]} — title+snippetにand_anyのいずれか1つが
#         含まれる場合のみ検知（ベースキーワード単体では検知しない）
_FRAUD_WORDS = ["偽物", "ニセモノ", "にせもの", "パチモン", "パチもん", "パチモノ",
                "コピー", "fake", "偽造品", "模倣品", "コピー商品", "模造品", "詐欺",
                "販売禁止商品", "規約違反", "強制返金",
                "中国"]

KEYWORDS = [
    # Qoo10 関連 — 複合クエリで検索精度を維持 + AND条件でスニペット検証
    {"query": "Qoo10 偽物",    "and_any": _FRAUD_WORDS},
    {"query": "Qoo10 ニセモノ", "and_any": _FRAUD_WORDS},
    {"query": "Qoo10 にせもの", "and_any": _FRAUD_WORDS},
    {"query": "Qoo10 パチモン", "and_any": _FRAUD_WORDS},
    {"query": "Qoo10 パチもん", "and_any": _FRAUD_WORDS},
    {"query": "Qoo10 パチモノ", "and_any": _FRAUD_WORDS},
    {"query": "Qoo10 コピー",   "and_any": _FRAUD_WORDS},
    {"query": "Qoo10 fake",    "and_any": _FRAUD_WORDS},
    {"query": "Qoo10 偽造品",   "and_any": _FRAUD_WORDS},
    {"query": "Qoo10 模倣品",   "and_any": _FRAUD_WORDS},
    {"query": "Qoo10 コピー商品", "and_any": _FRAUD_WORDS},
    {"query": "Qoo10 模造品",   "and_any": _FRAUD_WORDS},
    {"query": "Qoo10 詐欺",    "and_any": _FRAUD_WORDS},
    # メガ割 関連 — AND条件: 偽物系ワードと同時に存在する場合のみ検知
    {"query": "メガ割り", "and_any": _FRAUD_WORDS},
    {"query": "メガ割",   "and_any": _FRAUD_WORDS},
    # 規制・被害対応ワード — 実被害投稿の未捕捉パターンをカバー
    {"query": "Qoo10 販売禁止商品", "and_any": _FRAUD_WORDS},
    {"query": "Qoo10 規約違反",     "and_any": _FRAUD_WORDS},
    {"query": "Qoo10 強制返金",     "and_any": _FRAUD_WORDS},
]

# ── 환경변수 ────────────────────────────────────────────
SERPAPI_KEY       = os.getenv("SERPAPI_KEY", "")
SLACK_WEBHOOK_URL = os.getenv("SLACK_WEBHOOK_URL", "")
EMAIL_FROM        = os.getenv("EMAIL_FROM", "ragong@ebay.com")
EMAIL_TO_LIST     = [a.strip() for a in os.getenv("EMAIL_TO", "ragong@ebay.com").split(",")]
EMAIL_PASSWORD    = os.getenv("EMAIL_PASSWORD", "")

STREAMLIT_URL = "https://qoo10-monitor-kpcsgufhoixrfo6ekyxmc7.streamlit.app/"

RESULTS_DIR  = Path(__file__).parent / "results"
LOGS_DIR     = Path(__file__).parent / "logs"
HISTORY_FILE = RESULTS_DIR / "url_history.json"

# ── 판별 키워드 ─────────────────────────────────────────
REVIEW_SIGNALS = [
    "買った", "届いた", "購入した", "買ってしまった",
    "偽物が届いた", "ニセモノが届いた", "買わされた",
    "つかまされた", "やられた", "体験談", "実際に",
    "買ってみた", "被害", "詐欺にあった",
    "強制返金", "規約違反",
    "bought", "received", "purchased", "got a fake",
]
GUIDE_SIGNALS = [
    "見分け方", "見分け方法", "ガイド", "まとめ", "how to spot",
    "complete guide", "チェックポイント", "対策方法",
    "見極め", "注意点", "安全な買い方",
]
# 以下のワードが含まれる場合は常に除外（has_qoo10_ref 例外も適用しない）
EXCLUDE_KEYWORDS = [
    # ガイド・情報系
    "見分け方", "見分け方法",
    # 유료광고/PR 표시
    "#PR", "#広告", "#ad", "#sponsored", "#タイアップ", "#案件",
    # Qoo10 공식 약관/캠페인 보일러플레이트 (사용자 피해 신고문에는 나오지 않음)
    "弊社が判断した場合",
]
SOCIAL_DOMAINS = [
    "x.com", "twitter.com", "instagram.com", "tiktok.com",
    "note.com", "ameblo.jp", "youtube.com", "threads.net",
]

QOO10_URL_RE = re.compile(r"https?://(?:www\.)?qoo10\.jp/\S+")


# ── URL 履歴管理 ────────────────────────────────────────
def load_url_history() -> set:
    if not HISTORY_FILE.exists():
        return set()
    data = json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
    cutoff = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    return {url for url, date in data.items() if date >= cutoff}


def save_url_history(new_urls: dict[str, str]):
    existing: dict = {}
    if HISTORY_FILE.exists():
        existing = json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
    existing.update(new_urls)
    cutoff = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    pruned = {url: date for url, date in existing.items() if date >= cutoff}
    HISTORY_FILE.write_text(json.dumps(pruned, ensure_ascii=False, indent=2), encoding="utf-8")


# ── 検索 ───────────────────────────────────────────────
def search_keyword(keyword: str) -> list[dict]:
    from serpapi import GoogleSearch
    params = {
        "q":       keyword,
        "api_key": SERPAPI_KEY,
        "hl":      "ja",
        "gl":      "jp",
        "tbs":     "qdr:d2",  # 過去2日
        "num":     30,
    }
    return GoogleSearch(params).get_dict().get("organic_results", [])


# ── 危険度判定 ──────────────────────────────────────────
def classify(result: dict) -> str:
    text = (result.get("title", "") + " " + result.get("snippet", ""))
    url  = result.get("link", "")

    has_review = any(s in text for s in REVIEW_SIGNALS)
    has_guide  = any(s in text for s in GUIDE_SIGNALS)
    is_social  = any(d in url  for d in SOCIAL_DOMAINS)

    if has_review and not has_guide:
        return "HIGH"
    if has_review or is_social:
        return "MEDIUM"
    if has_guide:
        return "LOW"
    return "MEDIUM"


def proximity_ok(text: str, query: str, fraud_words: list[str], window: int = 100) -> bool:
    """アンカー語(Qoo10 / クエリ先頭語)とfake系ワードがwindow文字以内に共存するか確認。
    異なる文脈のテキストが混在するページ(知恵袋・TikTok等)の誤検知を防ぐ。
    """
    anchors = ["Qoo10"] if "Qoo10" in query else [query.split()[0], "Qoo10"]
    for anchor in anchors:
        pos = 0
        while True:
            idx = text.find(anchor, pos)
            if idx == -1:
                break
            nearby = text[max(0, idx - window): idx + len(anchor) + window]
            if any(fw in nearby for fw in fraud_words):
                return True
            pos = idx + 1
    return False


def extract_qoo10_link(result: dict) -> str:
    for field in ("link", "snippet", "title"):
        m = QOO10_URL_RE.search(result.get(field, ""))
        if m:
            return m.group(0)
    return "null"


# ── キーワードユーティリティ ────────────────────────────
def kw_query(kw) -> str:
    return kw if isinstance(kw, str) else kw["query"]

def kw_display(kw) -> str:
    return kw if isinstance(kw, str) else kw.get("display", kw["query"])

def kw_and_any(kw) -> list | None:
    return None if isinstance(kw, str) else kw.get("and_any")


# ── 検索実行 ────────────────────────────────────────────
def run_searches() -> list[dict]:
    date_str  = datetime.now().strftime("%Y-%m-%d")
    url_map:  dict[str, dict] = {}   # URL 기준 중복 제거 + 키워드 병합
    seen_urls = load_url_history()
    new_urls:  dict[str, str] = {}

    for kw in KEYWORDS:
        query   = kw_query(kw)
        display = kw_display(kw)
        and_any = kw_and_any(kw)

        print(f"  Searching: {query}" + (" [AND条件]" if and_any else ""))
        try:
            for r in search_keyword(query):
                url = r.get("link", "")

                # Qoo10 공식 페이지 제외 (PC·모바일 모두)
                if url.startswith("https://www.qoo10.jp/") or url.startswith("https://m.qoo10.jp/"):
                    continue

                # TikTok discover 집계 페이지 제외 (특정 게시물 아님)
                if "tiktok.com/discover/" in url:
                    continue

                # 이미 보고된 URL 중복 제외
                if url in seen_urls:
                    continue

                # Qoo10(またはクエリ語)とfake系ワードが100文字以内に共存するか確認
                text = r.get("title", "") + " " + r.get("snippet", "")
                if and_any and not proximity_ok(text, query, and_any):
                    continue

                # 見分け方/広告系は常に除外
                if any(w in text for w in EXCLUDE_KEYWORDS):
                    continue

                # 스니펫/타이틀에 Qoo10 URL이 포함된 외부 페이지는 포함
                has_qoo10_ref = bool(QOO10_URL_RE.search(
                    r.get("snippet", "") + " " + r.get("title", "")
                ))
                likelihood = classify(r)
                if likelihood == "LOW":
                    if not has_qoo10_ref:
                        continue
                    likelihood = "MEDIUM"

                if url in url_map:
                    # 동일 URL — 키워드 병합, 위험도는 높은 쪽 유지
                    entry = url_map[url]
                    if display not in entry["keyword"]:
                        entry["keyword"] += "・" + display
                    if likelihood == "HIGH":
                        entry["likelihood"] = "HIGH"
                else:
                    title = r.get("title", "")
                    url_map[url] = {
                        "date":       date_str,
                        "keyword":    display,
                        "url":        url,
                        "summary":    (title + " — " + r.get("snippet", ""))[:300],
                        "qoo10_link": extract_qoo10_link(r),
                        "likelihood": likelihood,
                        "search_url": "https://www.google.com/search?q=" + quote(title[:80]) + "&tbs=qdr:d2",
                    }
                    new_urls[url] = date_str
        except Exception as e:
            print(f"  [ERROR] {query}: {e}")

    RESULTS_DIR.mkdir(exist_ok=True)
    save_url_history(new_urls)
    rows = list(url_map.values())
    rows.sort(key=lambda x: 0 if x["likelihood"] == "HIGH" else 1)
    return rows


# ── Wayback Machine 自動保存 ────────────────────────────
def save_to_wayback(url: str) -> str:
    """Wayback Machine에 URL 저장 요청. 성공 시 스냅샷 URL, 실패 시 빈 문자열."""
    try:
        r = requests.get(
            f"https://web.archive.org/save/{url}",
            headers={"User-Agent": "Mozilla/5.0"},
            timeout=30,
            allow_redirects=True,
        )
        if r.status_code == 200 and "/web/" in r.url:
            return r.url
    except Exception as e:
        print(f"  [WARN] Wayback 저장 실패 ({url[:60]}): {e}")
    return ""


def save_wayback_batch(rows: list[dict]):
    """탐지된 URL 전체를 Wayback Machine에 저장하고 wayback_url 필드 업데이트."""
    import time
    print(f"  Wayback Machine 저장 중 ({len(rows)}건)...")
    for i, row in enumerate(rows):
        url = row.get("url", "")
        if not url:
            continue
        snapshot = save_to_wayback(url)
        row["wayback_url"] = snapshot
        status = snapshot if snapshot else "저장 실패"
        print(f"  [{i+1}/{len(rows)}] {status[:80]}")
        if i < len(rows) - 1:
            time.sleep(2)


# ── 번역 유틸리티 ───────────────────────────────────────
def _translate_batch_ja_ko(texts: list[str]) -> list[str]:
    """일본어 텍스트를 한국어로 배치 번역. 실패 시 원문 반환."""
    if not texts:
        return []
    try:
        from deep_translator import GoogleTranslator
        translator = GoogleTranslator(source="ja", target="ko")
        results = []
        for i in range(0, len(texts), 50):
            chunk = [t[:500] for t in texts[i:i+50]]
            translated = translator.translate_batch(chunk)
            results.extend(translated if translated else chunk)
        return results
    except Exception as e:
        print(f"  [WARN] 번역 실패: {e}")
        return texts


def _bilingual(jp: str, kr: str) -> str:
    """일본어+한국어 병기 셀 내용 생성"""
    kr = (kr or "").strip()
    if kr and kr != jp.strip():
        return f"{jp}\n{'─' * 22}\n[한국어]\n{kr}"
    return jp


# ── Excel作成 ───────────────────────────────────────────
def save_excel(rows: list[dict]) -> Path:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    RESULTS_DIR.mkdir(exist_ok=True)
    date_str = datetime.now().strftime("%Y%m%d")
    path = RESULTS_DIR / f"Qoo10_偽物レポート_{date_str}.xlsx"

    wb = openpyxl.Workbook()

    # ── 시트1: 상세 레포트 ──
    ws = wb.active
    ws.title = "위조품 레포트"

    headers = ["검색일 / 検索日", "검색 키워드", "URL", "개요 / 概要", "Qoo10 상품 / 商品P", "위험도 / 危険度", "검색확인 / 検索確認", "오탐지여부", "Status", "아카이브 / Archive"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="2F5496")
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    high_fill = PatternFill("solid", fgColor="FFD7D7")
    med_fill  = PatternFill("solid", fgColor="FFF3CD")

    print("  Translating summaries (JP→KR)...")
    kr_list = _translate_batch_ja_ko([r["summary"] for r in rows])

    for row, kr in zip(rows, kr_list):
        ws.append([
            row["date"], row["keyword"], row["url"],
            _bilingual(row["summary"], kr),
            row["qoo10_link"], row["likelihood"],
            "", "", "New",   # G=검색확인(placeholder), H=오탐지여부, I=Status
        ])
        fill = high_fill if row["likelihood"] == "HIGH" else med_fill
        for cell in ws[ws.max_row]:
            cell.fill = fill
        # G열: Google 검색결과 링크
        if row.get("search_url"):
            g_cell = ws.cell(row=ws.max_row, column=7)
            g_cell.value = "Google検索"
            g_cell.hyperlink = row["search_url"]
            g_cell.font = Font(bold=False, color="0563C1", underline="single")
            g_cell.fill = fill
        # J열: Wayback Machine 아카이브 링크 (저장된 스냅샷 URL 우선)
        j_cell = ws.cell(row=ws.max_row, column=10)
        j_cell.value = "Wayback"
        j_cell.hyperlink = row.get("wayback_url") or f"https://web.archive.org/web/{row['url']}"
        j_cell.font = Font(bold=False, color="0563C1", underline="single")
        j_cell.fill = fill

    for col, width in zip("ABCDEFGHIJ", [12, 22, 55, 70, 50, 10, 12, 12, 14, 10]):
        ws.column_dimensions[col].width = width

    # H열 헤더에 입력 안내 메모 추가
    from openpyxl.comments import Comment
    note = Comment("O = 오탐지 (오검지)\nX = 실검지 (위조품 확인)\n공백 = 미확인", "monitor")
    ws["H1"].comment = note

    # H열 O/X 드롭다운 (오탐지여부)
    dv_ox = DataValidation(type="list", formula1='"O,X"', allow_blank=True)
    ws.add_data_validation(dv_ox)
    dv_ox.add("H2:H10000")

    # I열 Status 드롭다운
    dv_status = DataValidation(type="list", formula1='"New,Reviewing,Actioned,Closed"', allow_blank=False)
    ws.add_data_validation(dv_status)
    dv_status.add("I2:I10000")

    ws.freeze_panes = "A2"

    # ── 시트2: 키워드별 집계 ──
    ws2 = wb.create_sheet("키워드별 집계")
    ws2.append(["검색 키워드", "HIGH 건수", "MEDIUM 건수", "합계 / 合計"])
    for cell in ws2[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    kw_high = Counter(r["keyword"] for r in rows if r["likelihood"] == "HIGH")
    kw_med  = Counter(r["keyword"] for r in rows if r["likelihood"] == "MEDIUM")

    for kw in KEYWORDS:
        label = kw_display(kw)
        h = kw_high.get(label, 0)
        m = kw_med.get(label, 0)
        row_cells = ws2.append([label, h, m, h + m])
        if h > 0:
            for cell in ws2[ws2.max_row]:
                cell.fill = high_fill
        elif m > 0:
            for cell in ws2[ws2.max_row]:
                cell.fill = med_fill

    ws2.column_dimensions["A"].width = 25
    for col in "BCD":
        ws2.column_dimensions[col].width = 12

    wb.save(path)
    print(f"  Excel saved: {path}")
    return path, kr_list


# ── メール送信 (Gmail SMTP) ─────────────────────────────
def send_email(path: Path, summary: str):
    import smtplib
    from email import encoders
    from email.mime.base import MIMEBase
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    if not EMAIL_PASSWORD:
        print("  [SKIP] EMAIL_PASSWORD not set.")
        return

    try:
        msg            = MIMEMultipart()
        msg["From"]    = EMAIL_FROM
        msg["To"]      = ", ".join(EMAIL_TO_LIST)
        msg["Subject"] = f"[Qoo10 위조품 모니터링 / 偽物モニタリング] {datetime.now().strftime('%Y-%m-%d')} 레포트"

        body = (
            f"Qoo10 위조품 모니터링 레포트입니다 / Qoo10 偽物モニタリングレポートです。\n\n"
            f"{summary}\n\n"
            f"첨부 Excel 파일을 확인해 주세요 / 添付のExcelファイルをご確認ください。\n\n"
            f"─────────────────────────────\n"
            f"📊 모니터링 대시보드 (실무자 공유용)\n"
            f"{STREAMLIT_URL}\n"
            f"오탐지여부 · Status 업데이트 후 💾 저장 버튼을 눌러주세요。\n"
            f"─────────────────────────────"
        )
        msg.attach(MIMEText(body, "plain", "utf-8"))

        with open(path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename={path.name}")
            msg.attach(part)

        with smtplib.SMTP("smtp.gmail.com", 587) as s:
            s.ehlo()
            s.starttls()
            s.login(EMAIL_FROM, EMAIL_PASSWORD)
            s.sendmail(EMAIL_FROM, EMAIL_TO_LIST, msg.as_string())
        print(f"  Email sent to {', '.join(EMAIL_TO_LIST)}")
    except Exception as e:
        print(f"  [ERROR] Email failed: {e}")


def send_no_issue_email():
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    if not EMAIL_PASSWORD:
        print("  [SKIP] EMAIL_PASSWORD not set.")
        return

    try:
        msg            = MIMEMultipart()
        msg["From"]    = EMAIL_FROM
        msg["To"]      = ", ".join(EMAIL_TO_LIST)
        msg["Subject"] = f"[Qoo10 위조품 모니터링 / 偽物モニタリング] {datetime.now().strftime('%Y-%m-%d')} 이상없음 / 異常なし"
        body = (
            f"오늘의 Qoo10 위조품 모니터링 결과, 새로운 검지가 없었습니다 / 本日のQoo10 偽物モニタリングの結果、新たな検知はありませんでした。\n\n"
            f"이상없음 / 異常なし ✅\n\n"
            f"─────────────────────────────\n"
            f"📊 모니터링 대시보드 (실무자 공유용)\n"
            f"{STREAMLIT_URL}\n"
            f"─────────────────────────────"
        )
        msg.attach(MIMEText(body, "plain", "utf-8"))

        with smtplib.SMTP("smtp.gmail.com", 587) as s:
            s.ehlo()
            s.starttls()
            s.login(EMAIL_FROM, EMAIL_PASSWORD)
            s.sendmail(EMAIL_FROM, EMAIL_TO_LIST, msg.as_string())
        print(f"  Email sent (異常なし) to {', '.join(EMAIL_TO_LIST)}")
    except Exception as e:
        print(f"  [ERROR] Email failed: {e}")


# ── Slack 通知 ──────────────────────────────────────────
def send_slack(summary: str, path: Path | None = None):
    if not SLACK_WEBHOOK_URL:
        print("  [SKIP] SLACK_WEBHOOK_URL not set.")
        return

    text = (
        f"*[Qoo10 偽物モニタリング]* {datetime.now().strftime('%Y-%m-%d')}\n"
        f"{summary}"
    )
    if path:
        text += f"\n📎 ファイル: `{path.name}` (添付メールも確認してください)"

    r = requests.post(SLACK_WEBHOOK_URL, json={"text": text}, timeout=10)
    r.raise_for_status()
    print("  Slack notification sent.")


# ── メイン ──────────────────────────────────────────────
def main():
    print(f"[{datetime.now():%Y-%m-%d %H:%M}] Qoo10 偽物モニタリング 開始")

    rows = run_searches()
    high = sum(1 for r in rows if r["likelihood"] == "HIGH")
    med  = sum(1 for r in rows if r["likelihood"] == "MEDIUM")
    summary = (
        f"  ・危険度 HIGH  : {high} 件\n"
        f"  ・危険度 MEDIUM: {med} 件\n"
        f"  ・合計         : {len(rows)} 件"
    )
    print(summary)

    if len(rows) == 0:
        send_no_issue_email()
        send_slack("✅ 本日は新たな検知なし（異常なし）")
    else:
        save_wayback_batch(rows)
        path, kr_list = save_excel(rows)
        send_email(path, summary)
        send_slack(summary, path)
        try:
            from gsheets import write_to_sheets
            write_to_sheets(rows, kr_list, "google")
        except Exception as e:
            print(f"  [WARN] Google Sheets 기록 실패 (Excel은 정상 저장됨): {e}")

    print(f"[{datetime.now():%Y-%m-%d %H:%M}] 完了")


if __name__ == "__main__":
    main()
